import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def inspect_excel_structure(file_path):
    """Detailed inspection of Excel file structure"""
    
    print(f"Inspecting: {file_path}")
    print("=" * 80)
    
    # First, let's see all sheet names
    xl_file = pd.ExcelFile(file_path)
    print(f"\nAvailable sheets: {xl_file.sheet_names}")
    print("=" * 80)
    
    # Open with openpyxl for more detailed inspection
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Sheets to inspect in detail
    sheets_to_inspect = ['Risk Map', 'Controls Mapping', 'Regulatory Requirements']
    
    for sheet_name in sheets_to_inspect:
        if sheet_name in wb.sheetnames:
            print(f"\n\n{'='*80}")
            print(f"SHEET: {sheet_name}")
            print(f"{'='*80}")
            
            sheet = wb[sheet_name]
            
            # Show first 10 rows with all non-empty cells
            print("\nFirst 10 rows (showing all non-empty cells):")
            print("-" * 80)
            
            for row_idx in range(1, 11):  # Excel rows are 1-indexed
                row_data = []
                for col_idx in range(1, 20):  # Check first 20 columns
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        col_letter = get_column_letter(col_idx)
                        row_data.append(f"{col_letter}: {cell.value}")
                
                if row_data:
                    print(f"Row {row_idx}: {' | '.join(row_data[:5])}")  # Show first 5 columns
                    if len(row_data) > 5:
                        print(f"         {' | '.join(row_data[5:10])}")  # Next 5 columns
                    if len(row_data) > 10:
                        print(f"         ... and {len(row_data)-10} more columns")
            
            # Try to identify header row
            print("\n" + "-" * 80)
            print("Attempting to identify header row...")
            
            # Check rows 1-5 for potential headers
            potential_headers = {}
            for row_idx in range(1, 6):
                non_empty_count = 0
                row_values = []
                for col_idx in range(1, 30):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        non_empty_count += 1
                        row_values.append(str(cell.value))
                
                if non_empty_count > 5:  # Likely a header row if it has many values
                    potential_headers[row_idx] = {
                        'count': non_empty_count,
                        'sample': row_values[:10]
                    }
            
            for row_idx, info in potential_headers.items():
                print(f"\nRow {row_idx} (potential header - {info['count']} columns):")
                print(f"  {info['sample']}")
            
            # Use pandas to read with different header rows
            print("\n" + "-" * 80)
            print("Reading with pandas (trying different header rows)...")
            
            for header_row in [0, 1, 2, 3]:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                    # Clean up column names
                    df.columns = [str(col).strip() for col in df.columns]
                    
                    print(f"\nWith header row {header_row}:")
                    print(f"  Shape: {df.shape}")
                    print(f"  Columns ({len(df.columns)}): {list(df.columns[:10])}")
                    if len(df.columns) > 10:
                        print(f"  ... and {len(df.columns)-10} more columns")
                    
                    # Show first few non-empty rows
                    non_empty_mask = df.notna().any(axis=1)
                    non_empty_df = df[non_empty_mask].head(3)
                    if not non_empty_df.empty:
                        print(f"  First non-empty row sample:")
                        for idx, row in non_empty_df.iterrows():
                            non_empty_cols = row.dropna()
                            if len(non_empty_cols) > 0:
                                print(f"    Row {idx}: {dict(list(non_empty_cols.items())[:5])}")
                                break
                    
                except Exception as e:
                    print(f"  Error with header row {header_row}: {e}")
    
    wb.close()

# Main execution
if __name__ == "__main__":
    file_path = "General AI Risk Map.xlsx"
    inspect_excel_structure(file_path)